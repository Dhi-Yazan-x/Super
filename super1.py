from tkinter import *
import math, random, os
import datetime
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook



class Super:
    def __init__(self,root):
        global End_name
        global End_bill
        global End_phone
        global x
        global fatora
        self.root=root
        self.root.geometry('1300x700+30+10')
        self.root.title('Super-Market: سوبر ماركت')
        self.root.resizable(False,False)
        self.root.iconbitmap('D:\\super\\icon.ico')
        title = Label(self.root,text='سوبر ماركت الكعدني يرحب بكم',fg='white',bg='#0B2F3A',font=('tajawal',15,'bold'))
        title.pack(fill=X)
        #=============== المتغيرات ===============
        
        #=========[data excel]==========
        self.wb=Workbook()
        self.ws=self.wb.active

        self.ws.title='customer'
        self.ws["A1"]='اسم العميل'
        self.ws["B1"]='رقم الهاتف'
        self.ws["C1"]='رقم الفاتورة'
        self.ws["D1"]='السعر الكلي'
        self.ws["E1"]='تاريخ البيع'
        self.wb.save('Dhi-yazan.xlsx')
        
        #==========[بقوليات :q1 --> q24 ]==========
        self.q1=IntVar()
        self.q2=IntVar()
        self.q3=IntVar()
        self.q4=IntVar()
        self.q5=IntVar()
        self.q6=IntVar()
        self.q7=IntVar()
        self.q8=IntVar()
        self.q9=IntVar()
        self.q10=IntVar()
        self.q11=IntVar()
        self.q12=IntVar()
        self.q13=IntVar()
        self.q14=IntVar()
        self.q15=IntVar()
        self.q16=IntVar()
        self.q17=IntVar()
        self.q18=IntVar()
        self.q19=IntVar()
        self.q20=IntVar()
        self.q21=IntVar()
        self.q22=IntVar()
        self.q23=IntVar()
        self.q24=IntVar()
        
        #==========[اللوازم المنزلية :qq1 --> qq24 ]==========
        self.qq1=IntVar()
        self.qq2=IntVar()
        self.qq3=IntVar()
        self.qq4=IntVar()
        self.qq5=IntVar()
        self.qq6=IntVar()
        self.qq7=IntVar()
        self.qq8=IntVar()
        self.qq9=IntVar()
        self.qq10=IntVar()
        self.qq11=IntVar()
        self.qq12=IntVar()
        self.qq13=IntVar()
        self.qq14=IntVar()
        self.qq15=IntVar()
        self.qq16=IntVar()
        self.qq17=IntVar()
        self.qq18=IntVar()
        self.qq19=IntVar()
        self.qq20=IntVar()
        self.qq21=IntVar()
        self.qq22=IntVar()
        self.qq23=IntVar()
        self.qq24=IntVar()

        #==========[ادوات كهربائية :qq1 --> qq24 ]==========
        self.qqq1=IntVar()
        self.qqq2=IntVar()
        self.qqq3=IntVar()
        self.qqq4=IntVar()
        self.qqq5=IntVar()
        self.qqq6=IntVar()
        self.qqq7=IntVar()
        self.qqq8=IntVar()
        self.qqq9=IntVar()
        self.qqq10=IntVar()
        self.qqq11=IntVar()
        self.qqq12=IntVar()
        self.qqq13=IntVar()
        self.qqq14=IntVar()
        self.qqq15=IntVar()
        self.qqq16=IntVar()
        self.qqq17=IntVar()
        self.qqq18=IntVar()
        self.qqq19=IntVar()
        self.qqq20=IntVar()
        self.qqq21=IntVar()
        self.qqq22=IntVar()
        self.qqq23=IntVar()
        self.qqq24=IntVar()
        #==========[متغيرات بيانات المشتري]==========
        self.now=datetime.datetime.now()
        self.date=self.now.strftime("%Y-%m-%d")
        self.x=datetime.datetime.now()
        self.namo = StringVar()
        self.phono = StringVar()
        self.fatora = StringVar()
        x=random.randint(1000,2000)
        self.fatora.set(str(x))

        #==========[متغيرات الحساب الكلي]==========
        self.bacoliat=StringVar()
        self.adoat=StringVar()
        self.kahraba=StringVar()
        
        #========== بيانات المستخدم ==========
        F1=Frame(root, bd=2,width=338,height=170,bg='#1254B9')
        F1.place(x=961,y=35)
        tit =Label(F1, text=': بيانات المشتري',font=('tajawal',13,'bold'),bg='#1254B9',fg='tomato')
        tit.place(x=185,y=0)
        his_name=Label(F1,text=' اسم المشتري ',font=('tajawal',10,'bold'),bg='#1254B9',fg='white')
        his_name.place(x=230,y=40)
        his_phone=Label(F1,text=' رقم المشتري ',font=('tajawal',10,'bold'),bg='#1254B9',fg='white')
        his_phone.place(x=235,y=70)
        his_num=Label(F1,text=' رقم الفاتورة ',font=('tajawal',10,'bold'),bg='#1254B9',fg='white')
        his_num.place(x=238,y=100)
        End_name=Entry(F1,textvariable=self.namo, justify='center')
        End_name.place(x=90,y=42)
        End_phone=Entry(F1,textvariable=self.phono, justify='center')
        End_phone.place(x=90,y=72)
        End_bill=Entry(F1,textvariable=self.fatora, justify='center')
        End_bill.place(x=90,y=102)
        btn_customer=Button(F1,text='بحث',font=('tajawal',10,'bold'),width=10,height=1,bg='#DBA901',cursor='hand2',command=self.search_invoice)
        btn_customer.place(x=3,y=42)
        
        
        btn_customer1=Button(F1,text='حذف فاتورة',font=('tajawal',10,'bold'),width=10,height=1,bg='#DBA901',cursor='hand2', command=self.delete_invoice)
        btn_customer1.place(x=3,y=90)
        #========== الفاتورة =========# عرض عنوان الفاتورة
        titdd = Label(F1, text='[ الفاتورة ]', font=('tajawal', 15, 'bold'), bg='#1254B9', fg='gold')
        titdd.place(x=125, y=135)

        # الإطار الذي سيحتوي على الفاتورة
        F3 = Frame(root, bd=2, width=338, height=399, bg='white')
        F3.place(x=961, y=205)

        # إضافة الـ Scrollbar والـ Text
        scort_y = Scrollbar(F3, orient=VERTICAL)
        self.textarea = Text(F3, yscrollcommand=scort_y.set)
        scort_y.pack(side=LEFT, fill=Y)
        scort_y.config(command=self.textarea.yview)
        self.textarea.pack(fill=BOTH, expand=1)
        #========== الاسعار ==========
        F4 = Frame (root, bd = 2 , width=657, height=112, bg='#1254B9')
        F4.place ( x = 641, y = 587 )
        hesab=Button (F4, text="الحساب", width=12, height= 1 , font=('tajawal',14,'bold'),bg='#DBA901',cursor='hand2',command=self.total)
        hesab.place (x = 520, y = 5 )
        fatora=Button (F4, text="تصدير الفاتورة", width=12, height= 1 , font=('tajawal',14,'bold'),bg='#DBA901',cursor='hand2',command=self.Errore)
        fatora.place (x = 520, y = 57 )
        clear=Button (F4, text="افراغ الحقول", width=12, height= 1 , font=('tajawal',14,'bold'),bg='#DBA901',cursor='hand2',command=self.clear)
        clear.place (x = 387, y = 5 )
        exite=Button (F4, text="اغلاق البرنامج", width=12, height= 1 , font=('tajawal',14,'bold'),bg='#DBA901',cursor='hand2',command=self.close)
        exite.place (x = 387, y = 57 )
        lblo1 = Label(F4,text='الحساب الكلي للبقوليات',font=('tajawal',11,'bold'),bg='#1254B9',fg='gold')
        lblo1.place(x=210,y=10)
        lblo2 = Label(F4,text=' حساب اللوازم المنزلية ',font=('tajawal',11,'bold'),bg='#1254B9',fg='gold')
        lblo2.place(x=217,y=40)
        lblo3 = Label(F4,text=' حساب ادوات الكهرباء ',font=('tajawal',11,'bold'),bg='#1254B9',fg='gold')
        lblo3.place(x=220,y=70)
        ento1 =Entry(F4,textvariable=self.bacoliat,width=24)
        ento1.place(x=40,y=14)
        ento2 =Entry(F4,textvariable=self.adoat,width=24)
        ento2.place(x=40,y=42)
        ento3 =Entry(F4,textvariable=self.kahraba,width=24)
        ento3.place(x=40,y=72)
        #========== القسم الاول للبرنامج ==========
        FF1=Frame(root,bd=2,width=318,height=664,bg='#1254B9')
        FF1.place(x=1,y=35)
        t=Label(FF1,text='البقوليات',font=('tajawal',14,'bold'),bg='#1254B9',fg='gold')
        t.place(x=122,y=0)
        bq1=Label(FF1,text=' رز عادي ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq1.place(x=230,y=50)
        bq2=Label(FF1,text=' رز بسمتي  ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq2.place(x=218,y=80)
        bq3=Label(FF1,text=' فاصولياء ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq3.place(x=215,y=110)
        bq4=Label(FF1,text=' عدس ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq4.place(x=232,y=140)
        bq5=Label(FF1,text=' معكرونة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq5.place(x=216,y=170)
        bq6=Label(FF1,text=' سكر ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq6.place(x=238,y=200)
        bq7=Label(FF1,text=' حمص ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq7.place(x=228,y=230)
        bq8=Label(FF1,text=' فول ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq8.place(x=240,y=270)
        bq9=Label(FF1,text=' الملح ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq9.place(x=232,y=300)
        bq10=Label(FF1,text='فلفل اسود  ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq10.place(x=202,y=330)
        bq11=Label(FF1,text=' فلفل احمر ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq11.place(x=205,y=370)
        bq12=Label(FF1,text=' القمح ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq12.place(x=230,y=400)
        bq13=Label(FF1,text=' الشعير ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq13.place(x=223,y=430)
        bq14=Label(FF1,text=' الشوفان ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq14.place(x=215,y=470)
        bq15=Label(FF1,text=' الذرة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq15.place(x=239,y=500)
        bq16=Label(FF1,text=' برغل ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq16.place(x=240,y=530)

        '''
        bq17=Label(FF1,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq17.place(x=250,y=570)
        bq18=Label(FF1,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq18.place(x=250,y=600)
        bq19=Label(FF1,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq19.place(x=250,y=630)
        bq20=Label(FF1,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq20.place(x=250,y=670)
        bq21=Label(FF1,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq21.place(x=250,y=700)
        bq22=Label(FF1,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq22.place(x=250,y=730)
        bq23=Label(FF1,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq23.place(x=250,y=770)
        bq24=Label(FF1,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq24.place(x=250,y=800)
        '''

        bqent1=Spinbox(FF1,from_=0,to_=100,textvariable=self.q1,width=12)
        bqent1.place(x=70,y=50)
        bqent2=Spinbox(FF1,from_=0,to_=100,textvariable=self.q2,width=12)
        bqent2.place(x=70,y=80)
        bqent3=Spinbox(FF1,from_=0,to_=100, textvariable=self.q3,width=12)
        bqent3.place(x=70,y=110)
        bqent4=Spinbox(FF1,from_=0,to_=100,textvariable=self.q4,width=12)
        bqent4.place(x=70,y=140)
        bqent5=Spinbox(FF1,from_=0,to_=100,textvariable=self.q5,width=12)
        bqent5.place(x=70,y=170)
        bqent6=Spinbox(FF1,from_=0,to_=100,textvariable=self.q6,width=12)
        bqent6.place(x=70,y=200)
        bqent7=Spinbox(FF1,from_=0,to_=100,textvariable=self.q7,width=12)
        bqent7.place(x=70,y=230)
        bqent8=Spinbox(FF1,from_=0,to_=100,textvariable=self.q8,width=12)
        bqent8.place(x=70,y=270)
        bqent9=Spinbox(FF1,from_=0,to_=100,textvariable=self.q9,width=12)
        bqent9.place(x=70,y=300)
        bqent10=Spinbox(FF1,from_=0,to_=100,textvariable=self.q10,width=12)
        bqent10.place(x=70,y=330)
        bqent11=Spinbox(FF1,from_=0,to_=100,textvariable=self.q11,width=12)
        bqent11.place(x=70,y=370)
        bqent12=Spinbox(FF1,from_=0,to_=100,textvariable=self.q12,width=12)
        bqent12.place(x=70,y=400)
        bqent13=Spinbox(FF1,from_=0,to_=100,textvariable=self.q13,width=12)
        bqent13.place(x=70,y=430)
        bqent14=Spinbox(FF1,from_=0,to_=100,textvariable=self.q14,width=12)
        bqent14.place(x=70,y=470)
        bqent15=Spinbox(FF1,from_=0,to_=100,textvariable=self.q15,width=12)
        bqent15.place(x=70,y=500)
        bqent16=Spinbox(FF1,from_=0,to_=100,textvariable=self.q16,width=12)
        bqent16.place(x=70,y=530)

        '''
        bqent17=Entry(FF1,width=12)
        bqent17.place(x=70,textvariable=self.q17,y=570)
        bqent18=Entry(FF1,width=12)
        bqent18.place(x=70,textvariable=self.q18,y=600)
        bqent19=Entry(FF1,width=12)
        bqent19.place(x=70,textvariable=self.q19,y=630)
        bqent20=Entry(FF1,width=12)
        bqent20.place(x=70,textvariable=self.q20,y=670)
        bqent21=Entry(FF1,width=12)
        bqent21.place(x=70,textvariable=self.q21,y=700)
        bqent22=Entry(FF1,width=12)
        bqent22.place(x=70,textvariable=self.q22,y=730)
        bqent23=Entry(FF1,width=12)
        bqent23.place(x=70,textvariable=self.q23,y=770)
        bqent24=Entry(FF1,width=12)
        bqent24.place(x=70,textvariable=self.q24,y=800)
        '''
        
        #========== القسم الثاني للبرنامج ==========
        FF2=Frame(root,bd=2,width=318,height=664,bg='#1254B9')
        FF2.place(x=321,y=35)
        t=Label(FF2,text='اللوازم المنزلية',font=('tajawal',14,'bold'),bg='#1254B9',fg='gold')
        t.place(x=105,y=0)
        bq1=Label(FF2,text=' صحن ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq1.place(x=250,y=50)
        bq2=Label(FF2,text=' مصفاة  ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq2.place(x=240,y=80)
        bq3=Label(FF2,text=' وعاء الخلط ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq3.place(x=215,y=110)
        bq4=Label(FF2,text=' كأس ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq4.place(x=250,y=140)
        bq5=Label(FF2,text=' سكين ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq5.place(x=246,y=170)
        bq6=Label(FF2,text=' إبريق ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq6.place(x=255,y=200)
        bq7=Label(FF2,text=' طنجرة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq7.place(x=243,y=230)
        bq8=Label(FF2,text=' سلة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq8.place(x=255,y=270)
        bq9=Label(FF2,text=' ملاعق ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq9.place(x=245,y=300)
        bq10=Label(FF2,text=' لوح تقطيع ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq10.place(x=217,y=330)
        bq11=Label(FF2,text=' فتاحة علب ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq11.place(x=220,y=370)
        bq12=Label(FF2,text=' مقشرة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq12.place(x=237,y=400)
        bq13=Label(FF2,text=' حفارة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq13.place(x=246,y=430)
        bq14=Label(FF2,text=' سلة قمامة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq14.place(x=215,y=470)
        bq15=Label(FF2,text=' اكياس ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq15.place(x=242,y=500)

        '''
        bq16=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq16.place(x=250,y=530)
        bq17=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq17.place(x=250,y=570)
        bq18=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq18.place(x=250,y=600)
        bq19=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq19.place(x=250,y=630)
        bq20=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq20.place(x=250,y=670)
        bq21=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq21.place(x=250,y=700)
        bq22=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq22.place(x=250,y=730)
        bq23=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq23.place(x=250,y=770)
        bq24=Label(FF2,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq24.place(x=250,y=800)
        '''
        
        bqent1=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq1,width=12)
        bqent1.place(x=70,y=50)
        bqent2=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq2,width=12)
        bqent2.place(x=70,y=80)
        bqent3=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq3,width=12)
        bqent3.place(x=70,y=110)
        bqent4=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq4,width=12)
        bqent4.place(x=70,y=140)
        bqent5=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq5,width=12)
        bqent5.place(x=70,y=170)
        bqent6=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq6,width=12)
        bqent6.place(x=70,y=200)
        bqent7=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq7,width=12)
        bqent7.place(x=70,y=230)
        bqent8=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq8,width=12)
        bqent8.place(x=70,y=270)
        bqent9=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq9,width=12)
        bqent9.place(x=70,y=300)
        bqent10=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq10,width=12)
        bqent10.place(x=70,y=330)
        bqent11=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq11,width=12)
        bqent11.place(x=70,y=370)
        bqent12=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq12,width=12)
        bqent12.place(x=70,y=400)
        bqent13=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq13,width=12)
        bqent13.place(x=70,y=430)
        bqent14=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq14,width=12)
        bqent14.place(x=70,y=470)
        bqent15=Spinbox(FF2,from_=0,to_=100,textvariable=self.qq15,width=12)
        bqent15.place(x=70,y=500)

        '''
        bqent16=Entry(FF2,width=12)
        bqent16.place(x=70,y=530)
        bqent17=Entry(FF2,width=12)
        bqent17.place(x=70,y=570)
        bqent18=Entry(FF2,width=12)
        bqent18.place(x=70,y=600)
        bqent19=Entry(FF2,width=12)
        bqent19.place(x=70,y=630)
        bqent20=Entry(FF2,width=12)
        bqent20.place(x=70,y=670)
        bqent21=Entry(FF2,width=12)
        bqent21.place(x=70,y=700)
        bqent22=Entry(FF2,width=12)
        bqent22.place(x=70,y=730)
        bqent23=Entry(FF2,width=12)
        bqent23.place(x=70,y=770)
        bqent24=Entry(FF2,width=12)
        bqent24.place(x=70,y=800)
        '''
        
        #========== القسم الثالث للبرنامج ==========
        FF3=Frame(root,bd=2,width=318,height=550,bg='#1254B9')
        FF3.place(x=641,y=35)
        t=Label(FF3,text='ادوات كهربائية',font=('tajawal',14,'bold'),bg='#1254B9',fg='gold')
        t.place(x=122,y=0)
        bq1=Label(FF3,text=' مكنسة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq1.place(x=230,y=50)
        bq2=Label(FF3,text=' تلفون  ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq2.place(x=240,y=80)
        bq3=Label(FF3,text=' مكرويف ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq3.place(x=230,y=110)
        bq4=Label(FF3,text=' خلاط ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq4.place(x=242,y=140)
        bq5=Label(FF3,text=' فرن غاز ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq5.place(x=230,y=170)
        bq6=Label(FF3,text=' مقلاة كهرباء ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq6.place(x=200,y=200)
        bq7=Label(FF3,text=' مروحة سقف ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq7.place(x=197,y=230)
        bq8=Label(FF3,text=' تلفزيون ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq8.place(x=235,y=270)
        bq9=Label(FF3,text=' فلتر ماء ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq9.place(x=229,y=300)
        bq10=Label(FF3,text=' غسالة اوتو',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq10.place(x=208,y=330)
        bq11=Label(FF3,text=' مكواة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq11.place(x=245,y=370)
        bq12=Label(FF3,text=' مبرد ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq12.place(x=250,y=400)
        bq13=Label(FF3,text=' توصيلة شحن ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq13.place(x=195,y=430)
        bq14=Label(FF3,text=' ثلاجة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq14.place(x=245,y=470)
        bq15=Label(FF3,text=' غسالة ',font=('tajawal',11,'bold'),bg='#1254B9',fg='white')
        bq15.place(x=239,y=500)

        '''
        bq16=Label(FF3,text=' جلاية صحون ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq16.place(x=250,y=530)
        bq17=Label(FF3,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq17.place(x=250,y=570)
        bq18=Label(FF3,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq18.place(x=250,y=600)
        bq19=Label(FF3,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq19.place(x=250,y=630)
        bq20=Label(FF3,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq20.place(x=250,y=670)
        bq21=Label(FF3,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq21.place(x=250,y=700)
        bq22=Label(FF3,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq22.place(x=250,y=730)
        bq23=Label(FF3,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq23.place(x=250,y=770)
        bq24=Label(FF3,text='  ',font=('tajawal',11),bg='#0B4C5F',fg='white')
        bq24.place(x=250,y=800)
        '''

        bqent1=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq1,width=12)
        bqent1.place(x=70,y=50)
        bqent2=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq2,width=12)
        bqent2.place(x=70,y=80)
        bqent3=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq3,width=12)
        bqent3.place(x=70,y=110)
        bqent4=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq4,width=12)
        bqent4.place(x=70,y=140)
        bqent5=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq5,width=12)
        bqent5.place(x=70,y=170)
        bqent6=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq6,width=12)
        bqent6.place(x=70,y=200)
        bqent7=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq7,width=12)
        bqent7.place(x=70,y=230)
        bqent8=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq8,width=12)
        bqent8.place(x=70,y=270)
        bqent9=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq9,width=12)
        bqent9.place(x=70,y=300)
        bqent10=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq10,width=12)
        bqent10.place(x=70,y=330)
        bqent11=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq11,width=12)
        bqent11.place(x=70,y=370)
        bqent12=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq12,width=12)
        bqent12.place(x=70,y=400)
        bqent13=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq13,width=12)
        bqent13.place(x=70,y=430)
        bqent14=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq14,width=12)
        bqent14.place(x=70,y=470)
        bqent15=Spinbox(FF3,from_=0,to_=100,textvariable=self.qqq15,width=12)
        bqent15.place(x=70,y=500)
        
        '''
        bqent16=Entry(FF3,width=12)
        bqent16.place(x=70,y=530)
        bqent17=Entry(FF3,width=12)
        bqent17.place(x=70,y=570)
        bqent18=Entry(FF3,width=12)
        bqent18.place(x=70,y=600)
        bqent19=Entry(FF3,width=12)
        bqent19.place(x=70,y=630)
        bqent20=Entry(FF3,width=12)
        bqent20.place(x=70,y=670)
        bqent21=Entry(FF3,width=12)
        bqent21.place(x=70,y=700)
        bqent22=Entry(FF3,width=12)
        bqent22.place(x=70,y=730)
        bqent23=Entry(FF3,width=12)
        bqent23.place(x=70,y=770)
        bqent24=Entry(FF3,width=12)
        bqent24.place(x=70,y=800)
        '''
        
        self.welcome()

    def total(self):
        self.rez=self.q1.get()*500
        self.reza=self.q2.get()*0.5
        self.fasoli=self.q3.get()*1
        self.ades=self.q4.get()*1.5
        self.makrona=self.q5.get()*2
        self.sakar=self.q6.get()*2
        self.hamas=self.q7.get()*1
        self.fol=self.q8.get()*1
        self.mlah=self.q9.get()*1.0
        self.flflasoed=self.q10.get()*1.5
        self.flflahmar=self.q11.get()*1
        self.qamh=self.q12.get()*2
        self.sheaer=self.q13.get()*1.5
        self.shofan=self.q14.get()*1
        self.drh=self.q15.get()*1
        self.borgel=self.q16.get()*1
        self.totalito=float(
            self.rez+
            self.reza+
            self.borgel+
            self.fasoli+
            self.ades+
            self.makrona+
            self.sakar+
            self.hamas+
            self.fol+
            self.mlah+
            self.flflasoed+
            self.flflahmar+
            self.qamh+
            self.sheaer+
            self.shofan+
            self.borgel

        )
        self.bacoliat.set(str(self.totalito)+" $ ")
        
        self.rez1=self.qq1.get()*500
        self.borgel1=self.qq2.get()*500
        self.fasoli1=self.qq3.get()*400
        self.ades1=self.qq4.get()
        self.makrona1=self.qq5.get()
        self.sakar1=self.qq6.get()
        self.hamas1=self.qq7.get()
        self.fol1=self.qq8.get()
        self.mlah1=self.qq9.get()
        self.flflasoed1=self.qq10.get()
        self.flflahmar1=self.qq11.get()
        self.qamh1=self.qq12.get()
        self.sheaer1=self.qq13.get()
        self.shofan1=self.qq14.get()
        self.kes1=self.qq15.get()
        self.khrba=float(
            self.rez1+
            self.borgel1+
            self.fasoli1+
            self.ades1+
            self.makrona1+
            self.sakar1+
            self.hamas1+
            self.fol1+
            self.mlah1+
            self.flflasoed1+
            self.flflahmar1+
            self.qamh1+
            self.sheaer1+
            self.shofan1+
            self.kes1

        )
        self.kahraba.set(str(self.khrba)+" $ ")

        self.rez2=self.qqq1.get()*500
        self.borgel2=self.qqq2.get()*500
        self.fasoli2=self.qqq3.get()*400
        self.ades2=self.qqq4.get()
        self.makrona2=self.qqq5.get()
        self.sakar2=self.qqq6.get()
        self.hamas2=self.qqq7.get()
        self.fol2=self.qqq8.get()
        self.mlah2=self.qqq9.get()
        self.flflasoed2=self.qqq10.get()
        self.flflahmar2=self.qqq11.get()
        self.qamh2=self.qqq12.get()
        self.sheaer2=self.qqq13.get()
        self.shofan2=self.qqq14.get()
        self.kes2=self.qqq15.get()
        self.adoatdd=float(
            self.rez2+
            self.borgel2+
            self.fasoli2+
            self.ades2+
            self.makrona2+
            self.sakar2+
            self.hamas2+
            self.fol2+
            self.mlah2+
            self.flflasoed2+
            self.flflahmar2+
            self.qamh2+
            self.sheaer2+
            self.shofan2+
            self.kes2

        )
        self.adoat.set(str(self.adoatdd)+" $ ")
        self.all=float(
            self.totalito+
            self.khrba+
            self.adoatdd
        )
        
        self.billing()


        
    def welcome(self):
        self.textarea.delete('1.0',END)
        self.textarea.insert(END,"   سوبر ماركت الكعدني للمواد الغذائيه ")
        self.textarea.insert(END,"\n========================================")
        self.textarea.insert(END,f"\n\t B.NUM    :{self.fatora.get()}")
        self.textarea.insert(END,f"\n\t NAME     :{self.namo.get()}")
        self.textarea.insert(END,f"\n\t PHONE    :{self.phono.get()}")
        self.textarea.insert(END,f"\n\t DATE     :{self.date}")
        self.textarea.insert(END,"\n========================================")
        self.textarea.insert(END,"\nالسعر\t        العدد\t        المشتريات")
        self.textarea.insert(END,"\n========================================")

        
    def billing(self):
            self.welcome()
            if self.q1.get()!=0:
                self.textarea.insert(END,f"\n {self.rez} ﷼\t\t{self.q1.get()}Kg\t       رز عادي")
            if self.q2.get()!=0:
                self.textarea.insert(END,f"\n {self.reza} ﷼\t\t{self.q2.get()}Kg\t      رز بسمتي")
            if self.q3.get()!=0:
                self.textarea.insert(END,f"\n {self.fasoli} ﷼\t\t{self.q3.get()}Kg\t      فاصولياء")
            if self.q4.get()!=0:
                self.textarea.insert(END,f"\n {self.ades} ﷼\t\t{self.q4.get()}Kg\t\t   عدس")
            if self.q5.get()!=0:
                self.textarea.insert(END,f"\n {self.makrona} ﷼\t\t{self.q5.get()}Kg\t       معكرونة")
            if self.q6.get()!=0:
                self.textarea.insert(END,f"\n {self.sakar} ﷼\t\t{self.q6.get()}Kg\t\t   سكر")
            if self.q7.get()!=0:
                self.textarea.insert(END,f"\n {self.hamas} ﷼\t\t{self.q7.get()}Kg\t\t   حمص")
            if self.q8.get()!=0:
                self.textarea.insert(END,f"\n {self.fol} ﷼\t\t{self.q8.get()}Kg\t\t   فول")
            if self.q9.get()!=0:
                self.textarea.insert(END,f"\n {self.mlah} ﷼\t\t{self.q9.get()}Kg\t\t   ملح")
            if self.q10.get()!=0:
                self.textarea.insert(END,f"\n {self.flflasoed} ﷼\t\t{self.q10.get()}Kg\t     فلفل اسود ")
            if self.q11.get()!=0:
                self.textarea.insert(END,f"\n {self.flflahmar} ﷼\t\t{self.q11.get()}Kg\t     فلفل احمر")
            if self.q12.get()!=0:
                self.textarea.insert(END,f"\n {self.qamh} ﷼\t\t{self.q12.get()}Kg\t\t   قمح")
            if self.q13.get()!=0:
                self.textarea.insert(END,f"\n {self.sheaer} ﷼\t\t{self.q13.get()}Kg\t\t  شعير")
            if self.q14.get()!=0:
                self.textarea.insert(END,f"\n {self.shofan} ﷼\t\t{self.q14.get()}Kg\t\t شوفان")
            if self.q15.get()!=0:
                self.textarea.insert(END,f"\n {self.drh} ﷼\t\t{self.q15.get()}Kg\t\t الذرة")
            if self.q16.get()!=0:
                self.textarea.insert(END,f"\n {self.borgel} ﷼\t\t{self.q16.get()}Kg\t\t  برغل")


            if self.qq1.get()!=0:
                self.textarea.insert(END,f"\n {self.rez1} ﷼\t\t{self.qq1.get()}\t\t   صحن")
            if self.qq2.get()!=0:
                self.textarea.insert(END,f"\n {self.borgel1} ﷼\t\t{self.qq2.get()}\t\t مصفاة")
            if self.qq3.get()!=0:
                self.textarea.insert(END,f"\n {self.fasoli1} ﷼\t\t{self.qq3.get()}\t      وعاء خلط")
            if self.qq4.get()!=0:
                self.textarea.insert(END,f"\n {self.ades1} ر\t\t{self.qq4.get()}\t\t   كأس")
            if self.qq5.get()!=0:
                self.textarea.insert(END,f"\n {self.makrona1} ﷼\t\t{self.qq5.get()}\t\t  سكين")
            if self.qq6.get()!=0:
                self.textarea.insert(END,f"\n {self.sakar1} ﷼\t\t{self.qq6.get()}\t\t  إبريق")
            if self.qq7.get()!=0:
                self.textarea.insert(END,f"\n {self.hamas1} ﷼\t\t{self.qq7.get()}\t\t  طنجرة")
            if self.qq8.get()!=0:
                self.textarea.insert(END,f"\n {self.fol1} ﷼\t\t{self.qq8.get()}\t\t   سلة")
            if self.qq9.get()!=0:
                self.textarea.insert(END,f"\n {self.mlah1} ﷼\t\t{self.qq9.get()}\t\t  ملاعق")
            if self.qq10.get()!=0:
                self.textarea.insert(END,f"\n {self.flflasoed1} ﷼\t\t{self.qq10.get()}\t      لوح تقطيع")
            if self.qq11.get()!=0:
                self.textarea.insert(END,f"\n {self.flflahmar1} ﷼\t\t{self.qq11.get()}\t      فتاحة علب")
            if self.qq12.get()!=0:
                self.textarea.insert(END,f"\n {self.qamh1} ﷼\t\t{self.qq12.get()}\t\t مقشرة")
            if self.qq13.get()!=0:
                self.textarea.insert(END,f"\n {self.sheaer1} ﷼\t\t{self.qq13.get()}\t\t حفارة")
            if self.qq14.get()!=0:
                self.textarea.insert(END,f"\n {self.shofan1} ﷼\t\t{self.qq14.get()}\t      سلة قمامة")
            if self.qq14.get()!=0:
                self.textarea.insert(END,f"\n {self.shofan2} ﷼\t\t{self.qq15.get()}\t\t اكياس")

            if self.qqq1.get()!=0:
                self.textarea.insert(END,f"\n {self.rez2} ﷼\t\t{self.qqq1.get()}\t\t مكنسة")
            if self.qqq2.get()!=0:
                self.textarea.insert(END,f"\n {self.borgel2} ﷼\t\t{self.qqq2.get()}\t\t تلفون")
            if self.qqq3.get()!=0:
                self.textarea.insert(END,f"\n {self.fasoli2} ﷼\t\t{self.qqq3.get()}\t\t مكرويف")
            if self.qqq4.get()!=0:
                self.textarea.insert(END,f"\n {self.ades2} ﷼\t\t{self.qqq4.get()}\t\t   خلاط")
            if self.qqq5.get()!=0:
                self.textarea.insert(END,f"\n {self.makrona2} ﷼\t\t{self.qqq5.get()}\t        فرن غاز")
            if self.qqq6.get()!=0:
                self.textarea.insert(END,f"\n {self.sakar2} ﷼\t\t{self.qqq6.get()}\t    مقلاة كهرباء")
            if self.qqq7.get()!=0:
                self.textarea.insert(END,f"\n {self.hamas2} ﷼\t\t{self.qqq7.get()}\t      مروحة سقف")
            if self.qqq8.get()!=0:
                self.textarea.insert(END,f"\n {self.fol2} ﷼\t\t{self.qqq8.get()}\t\t  تلفاز")
            if self.qqq9.get()!=0:
                self.textarea.insert(END,f"\n {self.mlah2} ﷼\t\t{self.qqq9.get()}\t       فلتر ماء")
            if self.qqq10.get()!=0:
                self.textarea.insert(END,f"\n {self.flflasoed2} ﷼\t\t{self.qqq10.get()}\t     غسالة اوتو ")
            if self.qqq11.get()!=0:
                self.textarea.insert(END,f"\n {self.flflahmar2} ﷼\t\t{self.qqq11.get()}\t\t  مكواة")
            if self.qqq12.get()!=0:
                self.textarea.insert(END,f"\n {self.qamh2} ﷼\t\t{self.qqq12.get()}\t\t  مبرد")
            if self.qqq13.get()!=0:
                self.textarea.insert(END,f"\n {self.sheaer2} ﷼\t\t{self.qqq13.get()}\t     توصيلة شحن")
            if self.qqq14.get()!=0:
                self.textarea.insert(END,f"\n {self.shofan2} ﷼\t\t{self.qqq14.get()}\t\t  ثلاجة")
            if self.qqq15.get()!=0:
                self.textarea.insert(END,f"\n {self.kes2} ﷼\t\t{self.qqq15.get()}\t\t  غسالة")
            self.textarea.insert(END,"\n----------------------------------------")
            self.textarea.insert(END,f"\n{self.all} ﷼\t\t\t المجموع الكلي  ")
            self.textarea.insert(END,"\n----------------------------------------")

    def clear(self):
        self.q1.set(0)
        self.q2.set(0)
        self.q3.set(0)
        self.q4.set(0)
        self.q5.set(0)
        self.q6.set(0)
        self.q7.set(0)
        self.q8.set(0)
        self.q9.set(0)
        self.q10.set(0)
        self.q11.set(0)
        self.q12.set(0)
        self.q13.set(0)
        self.q14.set(0)
        self.q15.set(0)
        self.q16.set(0)
        self.q17.set(0)
        self.q18.set(0)
        self.q19.set(0)
        self.q20.set(0)
        self.q21.set(0)
        self.q22.set(0)
        self.q23.set(0)
        self.q24.set(0)
        
        self.qq1.set(0)
        self.qq2.set(0)
        self.qq3.set(0)
        self.qq4.set(0)
        self.qq5.set(0)
        self.qq6.set(0)
        self.qq7.set(0)
        self.qq8.set(0)
        self.qq9.set(0)
        self.qq10.set(0)
        self.qq11.set(0)
        self.qq12.set(0)
        self.qq13.set(0)
        self.qq14.set(0)
        self.qq15.set(0)
        self.qq16.set(0)
        self.qq17.set(0)
        self.qq18.set(0)
        self.qq19.set(0)
        self.qq20.set(0)
        self.qq21.set(0)
        self.qq22.set(0)
        self.qq23.set(0)
        self.qq24.set(0)
        
        self.qqq1.set(0)
        self.qqq2.set(0)
        self.qqq3.set(0)
        self.qqq4.set(0)
        self.qqq5.set(0)
        self.qqq6.set(0)
        self.qqq7.set(0)
        self.qqq8.set(0)
        self.qqq9.set(0)
        self.qqq10.set(0)
        self.qqq11.set(0)
        self.qqq12.set(0)
        self.qqq13.set(0)
        self.qqq14.set(0)
        self.qqq15.set(0)
        self.qqq16.set(0)
        self.qqq17.set(0)
        self.qqq18.set(0)
        self.qqq19.set(0)
        self.qqq20.set(0)
        self.qqq21.set(0)
        self.qqq22.set(0)
        self.qqq23.set(0)
        self.qqq24.set(0)

        self.bacoliat.set('')
        self.adoat.set('')
        self.kahraba.set('')
        self.namo.set('')
        self.phono.set('')
        self.fatora.set('')


    def close(self):
        self.root.destroy()



    def search_invoice(self):
        search_query = self.fatora.get().strip()  

        if not search_query:
            messagebox.showwarning("تنبيه", "يرجى إدخال رقم الفاتورة للبحث.")
            return

        invoices_folder = 'D:\\super\\fatoras\\'
        found = False

        for file_name in os.listdir(invoices_folder):
            if file_name.endswith(".txt"): 
                with open(invoices_folder + file_name, 'r', encoding="utf-8") as file:
                    content = file.read()

                    if search_query in content:
                        found = True
                        self.show_invoice(file_name) 
                        break

        if not found:
            messagebox.showerror("لم يتم العثور على الفاتورة", "لم يتم العثور على الفاتورة باستخدام رقم الفاتورة المدخل.")
        
        self.search_bill()

    def delete_invoice(self):
        
        invoice_number = self.fatora.get() 
        invoice_file = f'D:\\super\\fatoras\\{invoice_number}.txt'

        if os.path.exists(invoice_file):
            try:
                os.remove(invoice_file)
                messagebox.showinfo("نجاح", f"تم حذف الفاتورة رقم {invoice_number} بنجاح.")
            except Exception as e:
                messagebox.showerror("خطأ", f"حدث خطأ أثناء حذف الفاتورة: {e}")
        else:
            messagebox.showerror("لم يتم العثور على الفاتورة", f"الفاتورة رقم {invoice_number} غير موجودة.")
        self.delete_bill()
        
    def show_invoice(self, file_name):
        invoices_folder = 'D:\\super\\fatoras\\'
        with open(invoices_folder + file_name, 'r', encoding="utf-8") as file:
            content = file.read()
            self.textarea.delete('1.0', END)
            self.textarea.insert(INSERT, content)



    def delete_bill(self):
        op = messagebox.askyesno("حذف", "هل تريد حذف بيانات العميل؟")
        if op > 0:
            self.delete_term = End_name.get() 
            if not self.delete_term:
                messagebox.showerror("خطأ", "يرجى إدخال اسم العميل للحذف")
                return

            excel = openpyxl.load_workbook('Dhi-yazan.xlsx')
            file = excel.active

            found = False
            rows = list(file.iter_rows(values_only=True))
            for index, row in enumerate(rows):
                if self.delete_term.lower() in str(row[0]).lower():
                    file.delete_rows(index + 1)
                    excel.save('Dhi-yazan.xlsx')
                    messagebox.showinfo("تم الحذف", "تم حذف بيانات العميل بنجاح")
                    found = True
                    break

            if not found:
                messagebox.showwarning("لم يتم العثور", "لم يتم العثور على اسم العميل للحذف")
                
        else:
            return

    def search_bill(self):
        self.search_term = End_name.get()
        if not self.search_term:
            messagebox.showerror("خطأ", "يرجى إدخال اسم للبحث")
            return
        
        excel = openpyxl.load_workbook('Dhi-yazan.xlsx')
        file = excel.active

        found = False
        for row in file.iter_rows(min_row=2, values_only=True):
            if self.search_term.lower() in str(row[0]).lower(): 
                messagebox.showinfo("بيانات العميل", f"البيانات:\nالاسم: {row[0]}\nرقم الهاتف: {row[1]}\n رقم الفاتوره: {row[2]}\nالإجمالي: {row[3]}\nالتاريخ: {row[4]}")
                found = True
                break
        
        if not found:
            messagebox.showwarning("غير موجود", "؟excel لم يتم العثور على بيانات العميل في ملف")
        
    def save1(self):

        op = messagebox.askyesno("حفظ", "هل تريد حفظ بيانات العميل؟")
        if op > 0:
            try:
                name = End_name.get()
                phone = End_phone.get()
                fator = End_bill.get()
                total = self.all
                datebuy =self.date

                excel = openpyxl.load_workbook('Dhi-yazan.xlsx')
                file = excel.active
                file.append([name, phone, fator, total, datebuy])
                excel.save('Dhi-yazan.xlsx')

                messagebox.showinfo("نجاح", "تم حفظ بيانات العميل بنجاح")
            except Exception as e:
                messagebox.showerror("خطأ", f"حدث خطأ أثناء الحفظ: {str(e)}")
        else:
            return


    def Errore(self):
        if self.namo.get =="" or self.phono.get()=="":
            messagebox.showerror("حدث خطأ","لا يجوز ترك حقل الاسم ورقم الهاتف فارغاً")
        else:
            if self.bacoliat.get()=="0.0 $" and self.adoat.get()=="0.0 $" and self.kahraba.get()=="0.0 $":
                messagebox.showerror("ليس هناك منتجات محددة ولم يتم اختيار احداها يجب اختيار عدد المنتجات","خطأ")
            else:
                if self.namo.get !="" or self.phono.get()!="":
                    self.savee()

    def savee(self):
        op = messagebox.askyesno("حفظ", "هل تريد حفظ الفاتورة؟")
        if op > 0:
            self.bb = self.textarea.get('1.0', END)

            try:
                with open('D:\\super\\fatoras\\' + str(self.fatora.get()) + ".txt", "w", encoding="utf-8") as f1:
                    f1.write(self.bb)

                messagebox.showinfo("تم الحفظ", f"تم حفظ الفاتورة رقم {self.fatora.get()} بنجاح.")
            except Exception as e:
                messagebox.showerror("خطأ", f"حدث خطأ أثناء حفظ الفاتورة: {e}")
        else:
            return
        
        self.save1()
        
root = Tk()
od = Super(root)
root.mainloop()